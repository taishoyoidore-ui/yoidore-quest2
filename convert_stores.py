import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# CSV読み込み
df = pd.read_csv('C:/AG_WORK/yoidore-quest2/STORES.csv', encoding='utf-8')

# excel出力
excel_path = 'C:/AG_WORK/yoidore-quest2/STORES.xlsx'
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='店舗一覧')

# openpyxlでスタイリング調整
wb = openpyxl.load_workbook(excel_path)
ws = wb['店舗一覧']

# ヘッダースタイル
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(name='Meiryo UI', size=11, bold=True, color='FFFFFF')
data_font = Font(name='Meiryo UI', size=10)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.font = data_font
        cell.border = thin_border
        col_letter = get_column_letter(cell.column)
        if col_letter in ['A', 'D', 'E', 'F', 'G', 'K', 'L']:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# 列幅調整
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        val_str = str(cell.value or '')
        length = sum(2 if ord(c) > 127 else 1 for c in val_str)
        if length > max_len:
            max_len = length
    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

# 1行目固定
ws.freeze_panes = 'A2'

wb.save(excel_path)
print("STORES.xlsx created successfully!")
