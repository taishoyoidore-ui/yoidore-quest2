import os
import sys
import re
import json
import openpyxl

VALID_DAYS = {'月', '火', '水', '木', '金', '土', '日'}
VALID_TYPES = {'はしご', '食事', 'ひと休み', '遊べる・エンタメ'}

def convert_excel_to_js():
    excel_path = os.path.join(os.path.dirname(__file__), 'STORES.xlsx')
    output_js_path = os.path.join(os.path.dirname(__file__), 'js', 'data.js')

    if not os.path.exists(excel_path):
        print(f"エラー: Excelファイルが見つかりません: {excel_path}")
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"エラー: Excelファイルの読み込みに失敗しました: {e}")
        sys.exit(1)

    errors = []

    # ------------------------------------------------------------------------
    # 1. ワークシート「イベント期間」の読み込み・バリデーション
    # ------------------------------------------------------------------------
    if 'イベント期間' not in wb.sheetnames:
        errors.append("エラー: ワークシート『イベント期間』が存在しません。")
        event_period = {"startDate": "", "endDate": ""}
    else:
        ws_event = wb['イベント期間']
        rows_event = list(ws_event.iter_rows(values_only=True))
        if len(rows_event) < 2:
            errors.append("エラー: ワークシート『イベント期間』に開始日・終了日のデータ行（2行目）が存在しません。")
            event_period = {"startDate": "", "endDate": ""}
        else:
            def parse_date_val(val):
                if val is None:
                    return ""
                if hasattr(val, 'strftime'):
                    return val.strftime('%Y-%m-%d')
                s = str(val).strip().split(' ')[0]
                return s

            start_str = parse_date_val(rows_event[1][0] if len(rows_event[1]) > 0 else None)
            end_str = parse_date_val(rows_event[1][1] if len(rows_event[1]) > 1 else None)

            date_regex = re.compile(r'^\d{4}-\d{2}-\d{2}$')
            if not date_regex.match(start_str):
                errors.append(f"エラー: ワークシート『イベント期間』の開始日 '{start_str}' の形式が不正です。'YYYY-MM-DD' 形式で指定してください。")
            if not date_regex.match(end_str):
                errors.append(f"エラー: ワークシート『イベント期間』の終了日 '{end_str}' の形式が不正です。'YYYY-MM-DD' 形式で指定してください。")

            event_period = {
                "startDate": start_str,
                "endDate": end_str
            }

    # ------------------------------------------------------------------------
    # 2. ワークシート「店舗一覧」の読み込み・バリデーション
    # ------------------------------------------------------------------------
    if '店舗一覧' not in wb.sheetnames:
        errors.append("エラー: ワークシート『店舗一覧』が存在しません。")
        ws_stores = None
        rows_stores = []
    else:
        ws_stores = wb['店舗一覧']
        rows_stores = list(ws_stores.iter_rows(values_only=True))

    if rows_stores:
        headers = [str(h or '').strip().lower() for h in rows_stores[0]]

        def get_col_idx(*names):
            for n in names:
                n_clean = n.strip().lower()
                if n_clean in headers:
                    return headers.index(n_clean)
            return -1

        id_idx = get_col_idx("id")
        name_idx = get_col_idx("店舗名", "名", "店名")
        area_idx = get_col_idx("エリア", "地域")
        category_idx = get_col_idx("カテゴリ", "カテゴリー", "ジャンル", "店の種類")
        type_idx = get_col_idx("酔いどれタイプ", "タイプ", "店舗タイプ", "スタイル", "セットタイプ")
        catchphrase_idx = get_col_idx("キャッチコピー", "コピー")

        set_title_idx = get_col_idx("酔いどれセット名", "セット名")
        set_content_idx = get_col_idx("セット内容", "内容")
        price_idx = get_col_idx("価格(円)", "価格", "金額", "セット価格(円)", "セット価格")
        charge_idx = get_col_idx("チャージ", "セットチャージ")

        days_idx = get_col_idx("提供日")
        hours_idx = get_col_idx("提供時間")
        limit_idx = get_col_idx("限定数")
        set_notes_idx = get_col_idx("セット備考", "備考・注意事項", "備考", "注意事項")

        quest_title_idx = get_col_idx("クエスト名")
        quest_price_idx = get_col_idx("クエスト価格(円)", "クエスト価格", "クエスト金額(円)", "クエスト金額")
        quest_charge_idx = get_col_idx("クエストチャージ")
        quest_content_idx = get_col_idx("クエスト内容", "イベント情報", "イベント")
        quest_notes_idx = get_col_idx("クエスト備考")

        payment_idx = get_col_idx("決済方法", "支払い方法")
        map_idx = get_col_idx("google map url", "googlemapurl", "マップurl")
        insta_idx = get_col_idx("instagram url", "instagramurl", "インスタurl")
        photo_idx = get_col_idx("photo", "写真", "画像")

        if name_idx == -1 or days_idx == -1 or hours_idx == -1:
            errors.append("エラー: 『店舗一覧』に必要な列（店舗名、提供日、提供時間）が見つかりません。")

        stores = []
        time_pattern = re.compile(r'^([0-1]?[0-9]|2[0-8]):([0-5][0-9])\s*-\s*([0-1]?[0-9]|2[0-8]):([0-5][0-9])$')

        for row_idx, row in enumerate(rows_stores[1:], start=2):
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue

            store_name = str(row[name_idx] or '').strip() if (name_idx != -1 and name_idx < len(row)) else ''
            if not store_name:
                continue

            store_id = str(row[id_idx] or '').strip() if (id_idx != -1 and id_idx < len(row) and row[id_idx]) else f"store-{str(row_idx-1).zfill(2)}"
            
            days_raw = str(row[days_idx] or '').strip() if (days_idx != -1 and days_idx < len(row)) else ''
            hours_raw = str(row[hours_idx] or '').strip() if (hours_idx != -1 and hours_idx < len(row)) else ''

            # --- 提供日チェック ---
            if not days_raw:
                errors.append(f"エラー: 店舗 '{store_name}' (行 {row_idx}): 『提供日』が空欄です。")
            else:
                day_list = [d.strip() for d in days_raw.split(',') if d.strip()]
                invalid_days = [d for d in day_list if d not in VALID_DAYS]
                if invalid_days or not day_list:
                    errors.append(
                        f"エラー: 店舗 '{store_name}' (行 {row_idx}): 『提供日』の指定 '{days_raw}' が不正です。"
                        f"カンマ区切りの曜日（例: 月,火,水,木,金,土,日）で指定してください。"
                    )
                else:
                    days_raw = ",".join(day_list)

            # --- 提供時間チェック ---
            if not hours_raw:
                errors.append(f"エラー: 店舗 '{store_name}' (行 {row_idx}): 『提供時間』が空欄です。")
            else:
                match = time_pattern.match(hours_raw)
                if not match:
                    errors.append(
                        f"エラー: 店舗 '{store_name}' (行 {row_idx}): 『提供時間』の指定 '{hours_raw}' が不正です。"
                        f"『HH:MM-HH:MM』（例: 17:00-23:00）形式で指定し、(L.O. 22:30) などの注記テキストは入れないでください。"
                    )
                else:
                    h1, m1, h2, m2 = match.groups()
                    hours_raw = f"{int(h1):02d}:{m1}-{int(h2):02d}:{m2}"

            # その他のフィールド読み込み（フォールバック補填なし・空欄は空欄のまま保持）
            area = str(row[area_idx] or '').strip() if (area_idx != -1 and area_idx < len(row) and row[area_idx]) else ''
            category = str(row[category_idx] or '').strip() if (category_idx != -1 and category_idx < len(row) and row[category_idx]) else ''
            store_type = str(row[type_idx] or '').strip() if (type_idx != -1 and type_idx < len(row) and row[type_idx]) else ''
            # 旧表記からの自動変換マッピング
            type_mapping = {
                'はしご向け': 'はしご',
                '食事向け': '食事',
                '休憩向け': 'ひと休み',
                'カフェ向け': 'ひと休み',
                'カフェ・ひと休み向け': 'ひと休み',
                'ひと休み向け': 'ひと休み',
                'バー・遊べる向け': '遊べる・エンタメ',
                '遊べる向け': '遊べる・エンタメ',
                'エンタメ向け': '遊べる・エンタメ'
            }
            if store_type in type_mapping:
                store_type = type_mapping[store_type]

            if store_type and store_type not in VALID_TYPES:
                errors.append(
                    f"エラー: 店舗 '{store_name}' (行 {row_idx}): 『酔いどれタイプ』の指定 '{store_type}' が不正です。"
                    f"指定可能なタイプは『はしご』『食事』『ひと休み』『遊べる・エンタメ』のいずれかです。"
                )
            catchphrase = str(row[catchphrase_idx] or '').strip() if (catchphrase_idx != -1 and catchphrase_idx < len(row) and row[catchphrase_idx]) else ''

            set_title = str(row[set_title_idx] or '').strip() if (set_title_idx != -1 and set_title_idx < len(row) and row[set_title_idx]) else ''
            set_content = str(row[set_content_idx] or '').strip() if (set_content_idx != -1 and set_content_idx < len(row) and row[set_content_idx]) else ''
            price_str = str(row[price_idx] or '').strip() if (price_idx != -1 and price_idx < len(row) and row[price_idx] is not None) else ''
            price_clean = ''.join(filter(str.isdigit, price_str))
            price = int(price_clean) if price_clean else 0

            charge_raw = str(row[charge_idx] or '').strip() if (charge_idx != -1 and charge_idx < len(row) and row[charge_idx]) else ''
            set_charge = '不要' if ('不要' in charge_raw or '無' in charge_raw) else '込'

            limit = str(row[limit_idx] or '').strip() if (limit_idx != -1 and limit_idx < len(row) and row[limit_idx]) else ''
            set_notes = str(row[set_notes_idx] or '').strip() if (set_notes_idx != -1 and set_notes_idx < len(row) and row[set_notes_idx]) else ''

            quest_title = str(row[quest_title_idx] or '').strip() if (quest_title_idx != -1 and quest_title_idx < len(row) and row[quest_title_idx]) else ''
            quest_price_str = str(row[quest_price_idx] or '').strip() if (quest_price_idx != -1 and quest_price_idx < len(row) and row[quest_price_idx] is not None) else ''
            quest_price_clean = ''.join(filter(str.isdigit, quest_price_str))
            quest_price = int(quest_price_clean) if quest_price_clean else 0

            quest_charge_raw = str(row[quest_charge_idx] or '').strip() if (quest_charge_idx != -1 and quest_charge_idx < len(row) and row[quest_charge_idx]) else ''
            quest_charge = '込' if ('込' in quest_charge_raw) else '不要'

            quest_content = str(row[quest_content_idx] or '').strip() if (quest_content_idx != -1 and quest_content_idx < len(row) and row[quest_content_idx]) else ''
            quest_notes = str(row[quest_notes_idx] or '').strip() if (quest_notes_idx != -1 and quest_notes_idx < len(row) and row[quest_notes_idx]) else ''
            is_quest_active = bool(quest_title or quest_content)

            payment_raw = str(row[payment_idx] or '').strip() if (payment_idx != -1 and payment_idx < len(row) and row[payment_idx]) else ''
            payment_methods = [p.strip() for p in re.split(r'[,/、\s]+', payment_raw) if p.strip()]

            g_map = str(row[map_idx] or '').strip() if (map_idx != -1 and map_idx < len(row) and row[map_idx]) else ''
            insta = str(row[insta_idx] or '').strip() if (insta_idx != -1 and insta_idx < len(row) and row[insta_idx]) else ''

            photo_file = str(row[photo_idx] or '').strip() if (photo_idx != -1 and photo_idx < len(row) and row[photo_idx]) else ''
            if photo_file:
                photo_file = re.sub(r'^photo[/\\]', '', photo_file, flags=re.IGNORECASE)
                photo_url = f"photo/{photo_file}"
            else:
                photo_url = None

            map_pos = {
                "top": f"{20 + (row_idx * 7) % 60}%",
                "left": f"{20 + (row_idx * 11) % 60}%"
            }

            store_obj = {
                "id": store_id,
                "name": store_name,
                "area": area,
                "category": category,
                "type": store_type,
                "isOpenToday": True,
                "isQuestActive": is_quest_active,
                "quest": {
                    "title": quest_title,
                    "price": quest_price,
                    "charge": quest_charge,
                    "content": quest_content,
                    "notes": quest_notes
                },
                "catchphrase": catchphrase,
                "yoidoreSet": {
                    "title": set_title,
                    "content": set_content,
                    "price": price,
                    "charge": set_charge,
                    "includeCharge": (set_charge == '込'),
                    "notes": set_notes
                },
                "conditions": {
                    "days": days_raw,
                    "hours": hours_raw,
                    "limit": limit,
                    "soldOutEnd": ("限定" in limit) or ("完売" in limit)
                },
                "paymentMethods": payment_methods,
                "googleMapUrl": g_map,
                "instagramUrl": insta,
                "photoUrl": photo_url,
                "mapPos": map_pos
            }
            stores.append(store_obj)

    # エラーが発生していた場合はメッセージを出力して終了
    if errors:
        print("\n".join(errors))
        print("\n変換処理を中止しました。STORES.xlsx を修正して再度実行してください。")
        sys.exit(1)

    areas_list = list(dict.fromkeys([s["area"] for s in stores if s["area"]]))
    categories_list = list(dict.fromkeys([s["category"] for s in stores if s["category"]]))
    types_list = list(dict.fromkeys([s["type"] for s in stores if s["type"]]))

    js_content = f"""/**
 * 大正酔いどれクエストⅡ - 店舗マスターデータ
 * ※このファイルは convert_xlsx_to_js.py によって STORES.xlsx から自動生成されたものです。
 * ※直接編集せず、STORES.xlsx を編集した後に update_data.bat を実行してください。
 */

const EVENT_PERIOD = {json.dumps(event_period, ensure_ascii=False, indent=2)};

const STORES_DATA = {json.dumps(stores, ensure_ascii=False, indent=2)};

const AREAS_LIST = {json.dumps(areas_list, ensure_ascii=False, indent=2)};
const CATEGORIES_LIST = {json.dumps(categories_list, ensure_ascii=False, indent=2)};
const TYPES_LIST = {json.dumps(types_list, ensure_ascii=False, indent=2)};

/**
 * 店舗が「今日・今」営業中かを動的に判定する関数
 */
function checkIsOpenToday(store, eventPeriod = (typeof EVENT_PERIOD !== 'undefined' ? EVENT_PERIOD : null), now = new Date()) {{
  if (!store) return false;

  // 1. イベント期間判定 (YYYY-MM-DD)
  if (eventPeriod && eventPeriod.startDate && eventPeriod.endDate) {{
    const y = now.getFullYear();
    const m = String(now.getMonth() + 1).padStart(2, '0');
    const d = String(now.getDate()).padStart(2, '0');
    const todayStr = `${{y}}-${{m}}-${{d}}`;

    if (todayStr < eventPeriod.startDate || todayStr > eventPeriod.endDate) {{
      return false;
    }}
  }}

  // 2. 曜日判定
  const dayNames = ['日', '月', '火', '水', '木', '金', '土'];
  const rawDays = (store.conditions && store.conditions.days) ? store.conditions.days : '';
  const allowedDays = rawDays.split(',').map(s => s.trim());

  // 3. 営業時間判定 (HH:MM-HH:MM)
  const rawHours = (store.conditions && store.conditions.hours) ? store.conditions.hours : '';
  if (!rawHours || !rawHours.includes('-')) {{
    return false;
  }}

  const parts = rawHours.split('-').map(s => s.trim());
  if (parts.length !== 2) return false;

  const [startStr, endStr] = parts;
  const [startH, startM] = startStr.split(':').map(Number);
  const [endH, endM] = endStr.split(':').map(Number);

  if (isNaN(startH) || isNaN(startM) || isNaN(endH) || isNaN(endM)) {{
    return false;
  }}

  const curMinutes = now.getHours() * 60 + now.getMinutes();
  const startMinutes = startH * 60 + startM;
  const endMinutes = endH * 60 + endM;

  const isOvernight = endMinutes <= startMinutes;

  if (!isOvernight) {{
    // 通常営業（同日内）
    const currentDayName = dayNames[now.getDay()];
    if (!allowedDays.includes(currentDayName)) return false;
    return curMinutes >= startMinutes && curMinutes < endMinutes;
  }} else {{
    // 深夜（日跨ぎ）営業
    if (curMinutes >= startMinutes) {{
      // 当日夜のシフト枠
      const currentDayName = dayNames[now.getDay()];
      return allowedDays.includes(currentDayName);
    }} else if (curMinutes < endMinutes) {{
      // 翌日早朝のシフト枠（前日営業枠の継続）
      const yesterday = new Date(now);
      yesterday.setDate(now.getDate() - 1);
      const yesterdayName = dayNames[yesterday.getDay()];
      return allowedDays.includes(yesterdayName);
    }} else {{
      return false;
    }}
  }}
}}

/**
 * Excel(XLSX) ArrayBufferをSTORES_DATAオブジェクト配列にパースする関数（Webサーバー閲覧時の動的更新用）
 */
function parseXLSXToStoresData(arrayBuffer) {{
  if (!arrayBuffer) return null;
  if (typeof XLSX === 'undefined') {{
    console.error('SheetJS (XLSX) ライブラリがロードされていません。');
    return null;
  }}

  const workbook = XLSX.read(arrayBuffer, {{ type: 'array' }});

  // イベント期間のパース
  let parsedEventPeriod = null;
  if (workbook.SheetNames.includes('イベント期間')) {{
    const eventSheet = workbook.Sheets['イベント期間'];
    const eventJson = XLSX.utils.sheet_to_json(eventSheet, {{ header: 1, defval: '' }});
    if (eventJson && eventJson.length >= 2) {{
      const startVal = String(eventJson[1][0] || '').trim().split(' ')[0];
      const endVal = String(eventJson[1][1] || '').trim().split(' ')[0];
      parsedEventPeriod = {{ startDate: startVal, endDate: endVal }};
      if (typeof EVENT_PERIOD !== 'undefined') {{
        EVENT_PERIOD.startDate = startVal;
        EVENT_PERIOD.endDate = endVal;
      }}
    }}
  }}

  const firstSheetName = workbook.SheetNames.includes('店舗一覧') ? '店舗一覧' : workbook.SheetNames[0];
  const worksheet = workbook.Sheets[firstSheetName];
  const jsonData = XLSX.utils.sheet_to_json(worksheet, {{ header: 1, defval: '' }});

  if (!jsonData || jsonData.length <= 1) return null;

  const headers = jsonData[0].map(h => String(h || '').trim());
  const newStores = [];

  for (let i = 1; i < jsonData.length; i++) {{
    const cols = jsonData[i].map(c => String(c === undefined || c === null ? '' : c).trim());
    if (cols.length === 0 || cols.every(c => c === '')) continue;

    const getVal = (...headerNames) => {{
      for (const name of headerNames) {{
        const target = String(name).toLowerCase().trim();
        const idx = headers.findIndex(h => String(h || '').toLowerCase().trim() === target);
        if (idx !== -1 && cols[idx] !== undefined && cols[idx] !== null && String(cols[idx]).trim() !== '') {{
          return String(cols[idx]).trim();
        }}
      }}
      return '';
    }};

    const id = getVal("ID", "id") || `store-${{String(i).padStart(2, '0')}}`;
    const name = getVal("店舗名", "名", "店名");
    if (!name) continue;

    const area = getVal("エリア", "地域");
    const category = getVal("カテゴリ", "カテゴリー", "ジャンル", "店の種類");
    const type = getVal("タイプ", "店舗タイプ", "スタイル");
    const catchphrase = getVal("キャッチコピー", "コピー");

    const set_title = getVal("酔いどれセット名", "セット名");
    const set_content = getVal("セット内容", "内容");
    const priceStr = getVal("価格(円)", "価格", "金額", "セット価格(円)", "セット価格").replace(/[^\\d]/g, '');
    const price = priceStr ? parseInt(priceStr, 10) : 0;
    const includeChargeStr = getVal("チャージ", "セットチャージ");
    const setCharge = (includeChargeStr.includes("不要") || includeChargeStr.includes("無")) ? "不要" : "込";

    const days = getVal("提供日");
    const hours = getVal("提供時間");
    const limit = getVal("限定数");
    const setNotes = getVal("セット備考", "備考・注意事項", "備考", "注意事項");

    const questTitle = getVal("クエスト名");
    const questPriceStr = getVal("クエスト価格(円)", "クエスト価格", "クエスト金額(円)", "クエスト金額").replace(/[^\\d]/g, '');
    const questPrice = questPriceStr ? parseInt(questPriceStr, 10) : 0;
    const questChargeStr = getVal("クエストチャージ");
    const questCharge = questChargeStr.includes("込") ? "込" : "不要";

    const questContent = getVal("クエスト内容", "イベント情報", "イベント");
    const questNotes = getVal("クエスト備考");
    const isQuestActive = !!(questTitle || questContent);

    const paymentMethodsRaw = getVal("決済方法", "支払い方法");
    const paymentMethods = paymentMethodsRaw
      ? paymentMethodsRaw.split(/[,/、\\s]+/).filter(Boolean)
      : [];

    const googleMapUrl = getVal("Google Map URL", "GoogleMapURL", "マップURL");
    const instagramUrl = getVal("Instagram URL", "InstagramURL", "インスタURL");
    let photoFileName = getVal("photo", "Photo", "PHOTO", "写真", "画像");
    if (photoFileName) {{
      photoFileName = photoFileName.replace(/^photo[/\\\\]/i, '');
    }}
    const photoUrl = photoFileName ? `photo/${{photoFileName}}` : null;

    const existing = (typeof STORES_DATA !== 'undefined') ? STORES_DATA.find(s => s.id === id || s.name === name) : null;
    const mapPos = existing ? existing.mapPos : {{ top: `${{20 + (i * 7) % 60}}%`, left: `${{20 + (i * 11) % 60}}%` }};

    const storeObj = {{
      id,
      name,
      area,
      category,
      type,
      isOpenToday: true,
      isQuestActive,
      quest: {{
        title: questTitle,
        price: questPrice,
        charge: questCharge,
        content: questContent,
        notes: questNotes
      }},
      catchphrase,
      yoidoreSet: {{
        title: set_title,
        content: set_content,
        price,
        charge: setCharge,
        includeCharge: setCharge === "込",
        notes: setNotes
      }},
      conditions: {{
        days,
        hours,
        limit,
        soldOutEnd: limit.includes("限定") || limit.includes("完売")
      }},
      paymentMethods,
      googleMapUrl,
      instagramUrl,
      photoUrl,
      mapPos
    }};

    storeObj.isOpenToday = checkIsOpenToday(storeObj, parsedEventPeriod || (typeof EVENT_PERIOD !== 'undefined' ? EVENT_PERIOD : null));
    newStores.push(storeObj);
  }}

  return newStores;
}}

function updateDataFromXLSX(arrayBuffer) {{
  const newStores = parseXLSXToStoresData(arrayBuffer);
  if (newStores && newStores.length > 0) {{
    STORES_DATA.length = 0;
    Array.prototype.push.apply(STORES_DATA, newStores);

    const areas = Array.from(new Set(STORES_DATA.map(s => s.area))).filter(Boolean);
    if (areas.length > 0) {{
      AREAS_LIST.length = 0;
      Array.prototype.push.apply(AREAS_LIST, areas);
    }}

    const categories = Array.from(new Set(STORES_DATA.map(s => s.category))).filter(Boolean);
    if (categories.length > 0) {{
      CATEGORIES_LIST.length = 0;
      Array.prototype.push.apply(CATEGORIES_LIST, categories);
    }}

    const types = Array.from(new Set(STORES_DATA.map(s => s.type))).filter(Boolean);
    if (types.length > 0) {{
      TYPES_LIST.length = 0;
      Array.prototype.push.apply(TYPES_LIST, types);
    }}

    return true;
  }}
  return false;
}}
"""

    with open(output_js_path, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f"成功: {len(stores)} 店舗のデータを {output_js_path} に出力しました。")
    return True

if __name__ == '__main__':
    convert_excel_to_js()
